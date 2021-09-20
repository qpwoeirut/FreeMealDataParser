from collections import defaultdict
from openpyxl import load_workbook

SHEET = load_workbook("frpm2021.xlsx", read_only=True, data_only=True)["FRPM School-Level Data "]
# yep, there's a trailing whitespace in the sheet name

DISTRICT_NAMES = [x[0] for x in SHEET.iter_rows(min_row=3, min_col=6, max_col=6, values_only=True)]
DISTRICTS = set(DISTRICT_NAMES)

with open("bay_area_districts.txt", 'r') as f:
    BAY_AREA_DISTRICTS = [s.replace('\xa0', ' ')
                           .replace("Schools", '')
                           .replace("School", '')
                           .replace('District', '')
                           .strip() for s in f.readlines()]
print(BAY_AREA_DISTRICTS)
BAD_PAIRS = [
    ("Campbell Union", "Campbell Union High"),
    ("San Francisco Unified", "South San Francisco Unified"),
    ("Los Altos", "Mountain View-Los Altos Union High")
]


def is_bad_pair(x: str, y: str) -> bool:
    return any([(x == pair[0] and y == pair[1]) or (x == pair[1] and y == pair[0]) for pair in BAD_PAIRS]) or \
           x == "Union Elementary" and y != "Union Elementary" or \
           y == "Union Elementary" and x != "Union Elementary"


def in_bay_area(district_name: str) -> str:
    for bay_area_district in BAY_AREA_DISTRICTS:
        if (district_name in bay_area_district or bay_area_district in district_name) and \
                not is_bad_pair(district_name, bay_area_district):
            return bay_area_district
    return ""


def identify_mismatches():
    remaining_districts = set(BAY_AREA_DISTRICTS)
    for district_name in DISTRICTS:
        x = in_bay_area(district_name)
        if x in remaining_districts:
            remaining_districts.remove(x)
    assert len(remaining_districts) == 0, remaining_districts


def main():
    k12_enrollment = defaultdict(lambda: 0)
    free_meal_count = defaultdict(lambda: 0)

    for district_name, data in zip(DISTRICT_NAMES,
                                   SHEET.iter_rows(min_row=3, min_col=18, max_col=19, values_only=True)):
        if data[0] > 0:  # some districts have 0 students apparently, which causes ZeroDivisionError
            k12_enrollment[district_name] += data[0]
            free_meal_count[district_name] += data[1]

    district_percentages = [
        (dn, free_meal_count[dn] * 100 / k12_enrollment[dn]) for dn in DISTRICTS
        if dn in k12_enrollment and k12_enrollment[dn] >= 100
    ]
    district_percentages.sort(key=lambda x: x[1], reverse=True)

    with open("district_free_meal_percentages.txt", 'w') as out:
        out.writelines([f"{str(percentage)[:5]},{district}\n" for district, percentage in district_percentages])

    with open("bay_area_sorted.txt", 'w') as out:
        out.writelines([
            f"{str(percentage)[:5]},{in_bay_area(district)},{district},{k12_enrollment[district]}\n" for district, percentage in district_percentages
            if in_bay_area(district) != ""
        ])


if __name__ == '__main__':
    identify_mismatches()
    main()
